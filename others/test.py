# create a main fumction using asyncio
import asyncio
from shutil import copyfile
from time import perf_counter
from icecream import ic
import csv
import datetime
from faker import Faker
import os
import asyncio
import aiofiles
import csv
import datetime
from faker import Faker
import os
import openpyxl
import pandas as pd
import polars as pl
import pandas as pd
import pandas as pd
import hashlib
from benedict import benedict
import pandas as pd
import pandas as pd
from openpyxl.styles import PatternFill


async def create_csv_file2(filename, delimiter, num_records):
    # Create the file name
    current_date = datetime.datetime.now().strftime("%Y%m%d")
    file_name = f"{filename}_{current_date}.csv"

    # Create the header lines
    header1 = f"FILENAME={file_name}"
    header2 = f"DATETIME={datetime.datetime.now().strftime('%Y%m%d%H%M%S')}"
    headers = [
        "id", "name", "address", "phone_number", "email", "date_of_birth", "job", "company", "credit_card_number", "points",
    ]
    av = ",".join([f"{s}" for s in headers])
    #ic(av)

    async with aiofiles.open(file_name, mode="w", newline="") as csv_file:
        writer = csv.writer(csv_file, quoting=csv.QUOTE_NONE)
        await writer.writerow([f"{header1} {header2}"])

    async with aiofiles.open(file_name, mode="a", newline="") as csv_file:
        writer = csv.writer(
            csv_file, delimiter=",", quotechar='"', quoting=csv.QUOTE_ALL
        )
        await writer.writerow(headers)

    # Write the data to the file
    async with aiofiles.open(file_name, mode="a", newline="") as csv_file:
        writer = csv.writer(
            csv_file, delimiter=delimiter, quotechar='"', quoting=csv.QUOTE_ALL
        )
        fake = Faker()
        for i in range(num_records):
            row = []
            row.append(i + 1)
            row.append(fake.name())
            row.append(
                fake.address()
            )  # Escape new line characters .replace('\n', '\\n')
            row.append(fake.phone_number())
            row.append(fake.email())
            row.append(fake.date_of_birth())
            row.append(fake.job())
            row.append(fake.company())
            row.append(fake.credit_card_number())
            row.append(fake.random_int())
            await writer.writerow(row)

    async with aiofiles.open(file_name, mode="a", newline="") as csv_file:
        writer = csv.writer(
            csv_file, delimiter=delimiter, quotechar='"', quoting=csv.QUOTE_NONE
        )
        await writer.writerow([f"RECORDCOUNT={num_records}"])

    print(f"File {file_name} created successfully!")
    return file_name


def find_mismatched_rows(df1, df2):
    """
    Compare two dataframes based on the 'idx' column and return rows where 'HashValue' is not matching.

    Parameters:
    df1 (pandas.DataFrame): First dataframe.
    df2 (pandas.DataFrame): Second dataframe.

    Returns:
    pandas.DataFrame: A dataframe containing the rows where 'HashValue' is not matching based on the 'idx' column.
    """
    merged_df = pd.merge(df1, df2, on="ConcatenatedKeys", suffixes=("_df1", "_df2"))
    mismatched_df = merged_df[merged_df["HashValue_df1"] != merged_df["HashValue_df2"]]
    return mismatched_df


def create_hash_dataframe(df, keys):
    """
    Create a separate dataframe with concatenated key columns and hash values of the complete record.

    Parameters:
    df (pandas.DataFrame): Input dataframe.
    keys (list): List of column names to use as keys for concatenation.

    Returns:
    pandas.DataFrame: A dataframe with two columns. The first column contains the concatenated key columns,
        and the second column contains the hash values of the complete record.
    """

    columnlist = df.columns.tolist()

    # Concatenate key columns
    df["ConcatenatedKeys"] = df[keys].astype(str).agg("▼".join, axis=1)

    # Calculate hash values
    df["HashValue"] = df.apply(
        lambda row: hashlib.sha256(row.to_string().encode()).hexdigest(), axis=1
    )
    # df['HashValue'] = df.apply(lambda row: hashlib.sha256(row.tostring().encode()).hexdigest(), axis=1, raw=True, result_type='reduce')

    # Create the separate dataframe
    hash_df = df[["ConcatenatedKeys", "HashValue", *columnlist]]

    return hash_df


def add_concatenated_keys(df, keys):
    """
    Add a column called 'ConcatenatedKeys' to the dataframe by concatenating the specified keys.

    Parameters:
    df (pandas.DataFrame): Input dataframe.
    keys (list): List of column names to use as keys for concatenation.

    Returns:
    pandas.DataFrame: The dataframe with the 'ConcatenatedKeys' column added.
    """
    df["ConcatenatedKeys"] = df[keys].astype(str).agg("▼".join, axis=1)


def get_extra_rows_using_hash(df1, df2):
    """
    Compare two dataframes and return extra rows based on the 'ConcatenatedKeys' column.

    Parameters:
    df1 (pandas.DataFrame): First dataframe.
    df2 (pandas.DataFrame): Second dataframe.

    Returns:
    tuple: A tuple containing two dataframes with extra rows from df1 and df2.
    """
    df1_extra = df1[~df1["ConcatenatedKeys"].isin(df2["ConcatenatedKeys"])]
    df2_extra = df2[~df2["ConcatenatedKeys"].isin(df1["ConcatenatedKeys"])]

    return df1_extra, df2_extra


def get_extra_rows_using_keys(df1, df2, keys):
    """
    Compare two dataframes based on specified keys.

    Parameters:
    df1 (pandas.DataFrame): First dataframe to compare.
    df2 (pandas.DataFrame): Second dataframe to compare.
    keys (list): List of column names to use as keys for comparison.

    Returns:
    tuple: A tuple containing two dataframes. The first dataframe contains extra records found in df1,
        and the second dataframe contains extra records found in df2.
    """
    # Set the keys as the index for both dataframes
    df1.set_index(keys, inplace=True)
    df2.set_index(keys, inplace=True)

    # Find extra records in df1
    extra_df1 = df1[~df1.index.isin(df2.index)]

    # Find extra records in df2
    extra_df2 = df2[~df2.index.isin(df1.index)]

    # Reset the index of both dataframes
    extra_df1.reset_index(inplace=True)
    extra_df2.reset_index(inplace=True)

    return extra_df1, extra_df2


def get_common_rows(df1, df2, keys):
    """
    Get the common rows between two dataframes based on specified keys.

    Parameters:
    df1 (pandas.DataFrame): First dataframe.
    df2 (pandas.DataFrame): Second dataframe.
    keys (list): List of column names to use as keys for comparison.

    Returns:
    pandas.DataFrame: A dataframe containing the common rows between df1 and df2.
    """
    # Merge the dataframes on the specified keys
    merged_df = pd.merge(df1, df2, on=keys)

    return merged_df


def get_mismatched_fields(df1, df2, mismatched_df):
    """
    Get the mismatched fields for each mismatched record between two dataframes.

    Parameters:
    df1 (pandas.DataFrame): First original dataframe.
    df2 (pandas.DataFrame): Second original dataframe.
    mismatched_df (pandas.DataFrame): Dataframe containing the mismatched rows.

    Returns:
    dict: A dictionary where the keys are the index of the mismatched records and the values are lists of mismatched fields.
    """
    mismatched_fields = {}

    for index, row in mismatched_df.iterrows():
        mismatched_fields[index] = []

        for column in mismatched_df.columns:
            if column in ["ConcatenatedKeys", "HashValue"]:
                if row[column] != df1.loc[index, column]:
                    mismatched_fields[index].append(column)

    return mismatched_fields


def compare_mismatched_rows_old(
    df1: pd.DataFrame, df2: pd.DataFrame, mismatched_df: pd.DataFrame
) -> benedict[str, benedict]:
    """
    Compare mismatched rows in two dataframes and return a dictionary with differing columns.

    Parameters:
    df1 (pandas.DataFrame): First dataframe.
    df2 (pandas.DataFrame): Second dataframe.
    mismatched_df (pandas.DataFrame): Dataframe with mismatched rows.

    Returns:
    benedict: A dictionary with column names as keys and dataframes as values.
    """
    result_dict = benedict()

    for index, row in mismatched_df.iterrows():
        keys = row["ConcatenatedKeys"]
        df1_row = df1[df1["ConcatenatedKeys"] == keys]
        df2_row = df2[df2["ConcatenatedKeys"] == keys]

        for column in df1.columns:
            if (
                column not in ["ConcatenatedKeys", "HashValue"]
                and df1_row[column].values[0] != df2_row[column].values[0]
            ):
                if column not in result_dict:
                    result_dict[column] = []
                result_dict[column].append(
                    {
                        "df1_value": df1_row[column].values[0],
                        "df2_value": df2_row[column].values[0],
                        "ConcatenatedKeys": keys,
                    }
                )

    return result_dict



def compare_mismatched_rows(
    df_left: pd.DataFrame, df_right: pd.DataFrame, mismatched_df: pd.DataFrame, df_left_extra : pd.DataFrame,df_right_extra : pd.DataFrame
) -> benedict[str, benedict]:
    """
    Compare mismatched rows in two dataframes and return a dictionary with differing columns.

    Parameters:
    df_left (pandas.DataFrame): First dataframe.
    df_right (pandas.DataFrame): Second dataframe.
    mismatched_df (pandas.DataFrame): Dataframe with mismatched rows.

    Returns:
    benedict: A dictionary with column names as keys and dataframes as values.
    """
    result_dict = benedict()

    for index, row in mismatched_df.iterrows():
        keys = row["ConcatenatedKeys"]
        df_left_row = df_left[df_left["ConcatenatedKeys"] == keys]
        df_right_row = df_right[df_right["ConcatenatedKeys"] == keys]

        for column in df_left.columns:
            if (
                column not in ["ConcatenatedKeys", "HashValue"]
                and df_left_row[column].values[0] != df_right_row[column].values[0]
            ):
                if column not in result_dict:
                    result_dict[column] = []
                result_dict[column].append(
                    {
                        "df_left_value": df_left_row[column].values[0],
                        "df_right_value": df_right_row[column].values[0],
                        "ConcatenatedKeys": keys,
                    }
                )
    for index, row in df_left_extra.iterrows():
        keys = row["ConcatenatedKeys"]
        df_left_row = df_left[df_left["ConcatenatedKeys"] == keys]
        for column in df_left.columns:
            if (
                column not in ["ConcatenatedKeys", "HashValue"]
            ):
                if column not in result_dict:
                    result_dict[column] = []
                result_dict[column].append(
                    {
                        "df_left_value": df_left_row[column].values[0],
                        "df_right_value": "Does_Not_Exists",
                        "ConcatenatedKeys": keys,
                    }
                )
    for index, row in df_right_extra.iterrows():
        keys = row["ConcatenatedKeys"]
        df_right_row = df_right[df_right["ConcatenatedKeys"] == keys]
        for column in df_right.columns:
            if (
                column not in ["ConcatenatedKeys", "HashValue"]
            ):
                if column not in result_dict:
                    result_dict[column] = []
                result_dict[column].append(
                    {
                        "df_left_value": "Does_Not_Exists",
                        "df_right_value": df_right_row[column].values[0],
                        "ConcatenatedKeys": keys,
                    }
                )

    return result_dict

def find_mismatched_rows(df_left, df_right):
    """
    Compare two dataframes based on the 'ConcatenatedKeys' column and return rows where 'HashValue' is not matching.

    Parameters:
    df_left (pandas.DataFrame): First dataframe.
    df_right (pandas.DataFrame): Second dataframe.

    Returns:
    pandas.DataFrame: A dataframe containing the rows where 'HashValue' is not matching based on the 'ConcatenatedKeys' column.
    """
    merged_df = pd.merge(df_left, df_right, on="ConcatenatedKeys", suffixes=("_df_left", "_df_right"))
    ic(merged_df)
    mismatched_df = merged_df[merged_df["HashValue_df_left"] != merged_df["HashValue_df_right"]]
    ic(mismatched_df)
    return mismatched_df

async def highlight_positive_cells(excel_filename, worksheet_name, column_name):
    """
    Highlights cells with positive values in the specified column of an Excel worksheet.

    Args:
        excel_filename (str): Path to the Excel file.
        worksheet_name (str): Name of the worksheet.
        column_name (str): Name of the column to check for positive values.

    Returns:
        None
    """
    try:
        # Read the Excel file
        df = pd.read_excel(excel_filename, sheet_name=worksheet_name)

        # Get the column index based on the column name
        column_index = df.columns.get_loc(column_name)

        # Create a workbook object
        wb = openpyxl.load_workbook(excel_filename)

        # Select the specified worksheet
        ws = wb[worksheet_name]

        # Define the red fill style
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

        # Iterate through rows and highlight positive cells
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=column_index + 1, max_col=column_index + 1):
            cell_value = row[0].value
            if cell_value is not None and cell_value > 0:
                row[0].fill = red_fill

        # Save the modified workbook
        wb.save(excel_filename)
        print(f"Positive values in column '{column_name}' highlighted in red.")

    except Exception as e:
        print(f"Error: {str(e)}")


async def highlight_cells_with_value(excel_filename, target_value):
    """
    Highlights cells in an Excel file that match the specified value.

    Args:
        excel_filename (str): Path to the Excel file.
        target_value (str): The value to search for in the cells.

    Returns:
        None
    """
    try:
        # Create a workbook object
        wb = openpyxl.load_workbook(excel_filename)

        # Iterate through all sheets in the workbook
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Define the red fill style
            red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

            # Iterate through all cells in the sheet
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value == target_value:
                        cell.fill = red_fill

        # Save the modified workbook
        wb.save(excel_filename)
        print(f"Cells with value '{target_value}' highlighted in red.")

    except Exception as e:
        print(f"Error: {str(e)}")


async def make_header_background_grey(excel_filename):
    """
    Modifies the background color of the header row in an Excel file to grey (if header cell has text).

    Args:
        excel_filename (str): Path to the Excel file.

    Returns:
        None
    """
    try:
        # Create a workbook object
        wb = openpyxl.load_workbook(excel_filename)

        # Iterate through all sheets in the workbook
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Define the grey fill style
            grey_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

            # Iterate through header cells (first row)
            for cell in ws[1]:
                if cell.value:
                    cell.fill = grey_fill

        # Save the modified workbook
        wb.save(excel_filename)
        print("Header background color set to grey for cells with text.")

    except Exception as e:
        print(f"Error: {str(e)}")



async def main():
    # task1 = asyncio.create_task(create_csv_file2("test1", ",", 500))
    # file_name = #ic(await task1)
    # copyfile(file_name, "test2_20240115.csv")

    df1 = pd.read_csv("/home/deck/devlopment/litestardemo/others/test1_20240115.csv", delimiter=",", quotechar='"', skiprows=1)
    df2 = pd.read_csv("/home/deck/devlopment/litestardemo/others/test2_20240115.csv", delimiter=",", quotechar='"', skiprows=1)

    # Remove last row from df1
    df1 = df1.iloc[:-1]

    # Remove last row from df2
    df2 = df2.iloc[:-1]

    # Specify datatype of first column to int
    df1["id"] = df1["id"].astype(int)
    df2["id"] = df2["id"].astype(int)
    hash_df1 = create_hash_dataframe(df1.copy(), ["id", "points"])
    hash_df2 = create_hash_dataframe(df2.copy(), ["id", "points"])

    extra_df1, extra_df2 = get_extra_rows_using_keys(
        df1.copy(), df2.copy(), ["id", "points"]
    )
    extra_df1, extra_df2 = get_extra_rows_using_hash(hash_df1, hash_df2)
    ic(extra_df1[:10])
    ic(extra_df2[:10])
    common_df = get_common_rows(hash_df1, hash_df2, ["ConcatenatedKeys"])
    
    mismatched_df = find_mismatched_rows(
        hash_df1[["ConcatenatedKeys", "HashValue"]].copy(),
        hash_df2[["ConcatenatedKeys", "HashValue"]].copy(),
    )
    ic(mismatched_df)

    #To DO this needs to be fulldf: there needs to be a way for unselecting batch, only one key column not working
    mismatched_dict = compare_mismatched_rows(hash_df1, hash_df2, mismatched_df)
    ic(mismatched_dict)



    # #ic(extra_df1)
    # #ic(extra_df2)
    # #ic(common_df.head(6))
    # #ic(mismatched_df)
    # #ic(mismatched_dict)
    #ic(len(common_df))
    #ic(len(df1))
    #ic(len(df1))
    #ic(len(mismatched_df))
    # for column_name, differ_items in mismatched_dict.items():
    #     for bdt in differ_items:
    #         #ic(bdt.ConcatenatedKeys, bdt.df1_value, bdt.df2_value)

    # Create a new Excel writer object
    common_df.set_index('ConcatenatedKeys', inplace=False)
    passed_col_df = pd.DataFrame(columns=['column_name', 'ConcatenatedKeys', 'value_df1', 'value_df2'])
    #ic(passed_col_df)


    passed_benedict_list : list[benedict] = []

    for column in df1.columns:
        #ic(column)
        for index, common_row in common_df.head(10).iterrows():
                    # New row as a DataFrame
            item =  benedict()
            item.column_name = column
            item.ConcatenatedKeys = common_row.ConcatenatedKeys
            item.value_df1 = common_row[f'{column}_x']
            item.value_df2 = common_row[f'{column}_y']
            passed_benedict_list.append(item)
    passed_col_df = pd.DataFrame(passed_benedict_list)

        











    with pd.ExcelWriter("output.xlsx", engine="xlsxwriter") as writer:
        # Create a new dataframe with the column names from df1
        summary_df = pd.DataFrame({"ColumnName": df1.columns})

        # Write the dataframe to the Excel sheet
        summary_df["MatchedRowsCount"] = summary_df["ColumnName"].apply(
            lambda col: len(common_df)
            - (len(mismatched_dict[col]) if col in mismatched_dict else 0)
        )
        summary_df["MisMatchedRowsCount"] = summary_df["ColumnName"].apply(
            lambda col: (len(mismatched_dict[col]) if col in mismatched_dict else 0)
        )

        summary_df["ExtraInFirst"] = summary_df["ColumnName"].apply(
            lambda col: len(extra_df1)
        )
        summary_df["ExtraInSecond"] = summary_df["ColumnName"].apply(
            lambda col: len(extra_df2)
        )

        summary_df["TotalRowsInFirst"] = len(df1)
        summary_df["TotalRowsInSecond"] = len(df2)

        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        passed_col_df.to_excel(writer, sheet_name="PassRecordSummary", index=False)


        for column_name, differ_items in mismatched_dict.items():
            mismatched_df = pd.DataFrame(columns=['ConcatenatedKeys', 'value_df1', 'value_df2'])
            mismatched_bendict_list : list[benedict] = []
            for dit in differ_items:
                item =  benedict()
                item.ConcatenatedKeys = dit.ConcatenatedKeys
                item.value_df1 = dit.df1_value
                item.value_df2 = dit.df2_value
                mismatched_bendict_list.append(item)
            mismatched_df = pd.DataFrame(mismatched_bendict_list)
            mismatched_df.to_excel(writer, sheet_name=f"{column_name}_miss", index=False)




# run the main function only from command line
if __name__ == "__main__":
    a = perf_counter()
    asyncio.run(main())
    b = perf_counter()
    #ic(b - a)


