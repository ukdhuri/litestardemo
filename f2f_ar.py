import random
import shutil
from typing import List
import uuid
import pandas as pd
from icecream import ic
from faker import Faker
from datetime import datetime, timedelta,time
import pendulum
from pathlib import Path
import pyarrow as pa
import os
import pyarrow.csv as csv
import io 
import subprocess
from tabulate import tabulate
import time as tx
from benedict import benedict
import openpyxl
from openpyxl.styles import PatternFill
import hashlib
import duckdb
import duckdb
import pyarrow.parquet as pq

def read_config(file_path,filename1,filename2):
    # Read the Excel file
    excel_data = pd.read_excel(file_path, sheet_name=None)
    # Extract the necessary information from the sheets
    meta1 = excel_data['meta1']
    meta2 = excel_data['meta2']
    colmapping = excel_data['colmapping']
    # Get the FILE_PATH from meta1 and meta2
    file_path1 = meta1.loc[meta1['MKey'] == 'FILE_PATH', 'MValue'].values[0]
    file_path2 = meta2.loc[meta2['MKey'] == 'FILE_PATH', 'MValue'].values[0]
    ic(meta1)
    ic(meta2)
    ic(colmapping)
    # Create the complete file names
    filename1 = Path(file_path1,filename1) 
    filename2 = Path(file_path2,filename2)
    filename_l = f'{filename1}.left'
    filename_r = f'{filename2}.right'

    shutil.copy2(filename1, filename_l)
    shutil.copy2(filename2, filename_r)

    
     
    return colmapping,meta1,meta2,filename_l,filename_r

class StreamWrapper():
    
    def __init__(self, obj=None):
        self.delegation = obj
        self.header = None

    def __getattr__(self, *args, **kwargs):
        # any other call is delegated to the stream/object
        return(self.delegation.__getattribute__(*args, **kwargs))

    def read(self, size=None, *args, **kwargs):
        bytedata = self.delegation.read(size, *args, **kwargs)          
        # .. the rest of the logic pre-processes the byte data, 
        # identifies the header and number of columns (which are retained persistently),
        # and then strips out extra columns when encountered
        return(bytedata)



def find_mismatched_rows(df_left, df_right):
    """
    Compare two dataframes based on the 'ConcatenatedKeys' column and return rows where 'HashValue' is not matching.

    Parameters:
    df_left (pandas.DataFrame): First dataframe.
    df_right (pandas.DataFrame): Second dataframe.

    Returns:
    pandas.DataFrame: A dataframe containing the rows where 'HashValue' is not matching based on the 'ConcatenatedKeys' column.
    """
    merged_df = pd.merge(df_left, df_right, on="ConcatenatedKeys", suffixes=("_df_left", "_df_right"))
    ic(merged_df)
    mismatched_df = merged_df[merged_df["HashValue_df_left"] != merged_df["HashValue_df_right"]]
    ic(mismatched_df)
    return mismatched_df

def create_hash_dataframe(df, keys):
    """
    Create a separate dataframe with concatenated key columns and hash values of the complete record.

    Parameters:
    df (pandas.DataFrame): Input dataframe.
    keys (list): List of column names to use as keys for concatenation.

    Returns:
    pandas.DataFrame: A dataframe with two columns. The first column contains the concatenated key columns,
        and the second column contains the hash values of the complete record.
    """

    columnlist = df.columns.tolist()

    # Concatenate key columns
    #df["ConcatenatedKeys"] = df[keys].astype(str).agg("▼".join, axis=1)

    df = duckdb.query(f"SELECT md5(concat_ws('',{','.join(str(item) for item in df.columns)})) as HashValue, * FROM df").to_df()
    print(df)
    df = duckdb.query(f"SELECT concat_ws('▼',{','.join(str(item) for item in keys)}) as ConcatenatedKeys, * FROM df").to_df()
    print(df)


    # Calculate hash values
    # df["HashValue"] = df.apply(
    #     lambda row: hashlib.sha256(row.to_string().encode()).hexdigest(), axis=1
    # )
    # df['HashValue'] = df.apply(lambda row: hashlib.sha256(row.tostring().encode()).hexdigest(), axis=1, raw=True, result_type='reduce')

    # Create the separate dataframe
    hash_df = df[["ConcatenatedKeys", "HashValue", *columnlist]]

    return hash_df

def create_hash_dataframe_duckdb(file_name, keys, columnlist, outputfile_name):
    """
    Create a separate dataframe with concatenated key columns and hash values of the complete record.

    Parameters:
    df (pandas.DataFrame): Input dataframe.
    keys (list): List of column names to use as keys for concatenation.

    Returns:
    pandas.DataFrame: A dataframe with two columns. The first column contains the concatenated key columns,
        and the second column contains the hash values of the complete record.
    """ 


    # Concatenate key columns
    #df["ConcatenatedKeys"] = df[keys].astype(str).agg("▼".join, axis=1)
    #query = "SELECT concat_ws('▼',{','.join(str(item) for item in keys)}) as ConcatenatedKeys, md5(concat_ws('',{','.join(str(item) for item in columnlist)})) as HashValue, * FROM '{file_name}' as spf"
    #query = 'SELECT 42'
    #df = duckdb.query(f"SELECT concat_ws('▼',{','.join(str(item) for item in keys)}) as ConcatenatedKeys, md5(concat_ws('',{','.join(str(item) for item in columnlist)})) as HashValue, * FROM '{file_name}' as spf").to_df()
    duckdb.query(f"SELECT concat_ws('▼',{','.join(str(item) for item in keys)}) as ConcatenatedKeys, md5(concat_ws('',{','.join(str(item) for item in columnlist)})) as HashValue FROM '{file_name}' as spf").to_parquet(file_name=outputfile_name,compression='gzip')
    #duckdb.sql(f"COPY ({query}) TO '/home/deck/Documents/inbound/sscc.parquet'")      # Copy to a Parquet file




def add_concatenated_keys(df, keys):
    """
    Add a column called 'ConcatenatedKeys' to the dataframe by concatenating the specified keys.

    Parameters:
    df (pandas.DataFrame): Input dataframe.
    keys (list): List of column names to use as keys for concatenation.

    Returns:
    pandas.DataFrame: The dataframe with the 'ConcatenatedKeys' column added.
    """
    df["ConcatenatedKeys"] = df[keys].astype(str).agg("▼".join, axis=1)



def get_extra_rows_using_hash(df_left, df_right):
    """
    Compare two dataframes and return extra rows based on the 'ConcatenatedKeys' column.

    Parameters:
    df_left (pandas.DataFrame): First dataframe.
    df_right (pandas.DataFrame): Second dataframe.

    Returns:
    tuple: A tuple containing two dataframes with extra rows from df_left and df_right.
    """
    df_left_extra = df_left[~df_left["ConcatenatedKeys"].isin(df_right["ConcatenatedKeys"])]
    df_right_extra = df_right[~df_right["ConcatenatedKeys"].isin(df_left["ConcatenatedKeys"])]

    return df_left_extra, df_right_extra


def get_extra_rows_using_keys(df1, df2, keys):
    """
    Compare two dataframes based on specified keys.

    Parameters:
    df1 (pandas.DataFrame): First dataframe to compare.
    df2 (pandas.DataFrame): Second dataframe to compare.
    keys (list): List of column names to use as keys for comparison.

    Returns:
    tuple: A tuple containing two dataframes. The first dataframe contains extra records found in df1,
        and the second dataframe contains extra records found in df2.
    """
    # Set the keys as the index for both dataframes
    df1.set_index(keys, inplace=True)
    df2.set_index(keys, inplace=True)

    # Find extra records in df1
    extra_df1 = df1[~df1.index.isin(df2.index)]

    # Find extra records in df2
    extra_df2 = df2[~df2.index.isin(df1.index)]

    # Reset the index of both dataframes
    extra_df1.reset_index(inplace=True)
    extra_df2.reset_index(inplace=True)

    return extra_df1, extra_df2

def get_extra_rows_using_keys_duckdb(filename_l, filename_r, outputfile_l, outputfile_r, datafile_l, datafile_r, keys, compare_row_limit=5000):
    query = f"""
    SELECT *
    FROM (
    SELECT ConcatenatedKeys FROM '{filename_l}'
    EXCEPT
    SELECT ConcatenatedKeys FROM '{filename_r}'
    ) AS extra_rows_in_a;
    """
    duckdb.sql(query).to_parquet(outputfile_l,compression='gzip')
    query = f"""
    SELECT *
    FROM (
    SELECT ConcatenatedKeys FROM '{filename_r}'
    EXCEPT
    SELECT ConcatenatedKeys FROM '{filename_l}'
    ) AS extra_rows_in_a;
    """
    duckdb.sql(query).to_parquet(outputfile_r,compression='gzip')
    extra_df_l_sql = f"""
    SELECT concat_main.* from (SELECT concat_ws('▼',{','.join(str(item) for item in keys)}) as ConcatenatedKeys,* from '{datafile_l}') AS concat_main JOIN '{outputfile_l}' extra_l on concat_main.ConcatenatedKeys = extra_l.ConcatenatedKeys limit {compare_row_limit}
    """
    extra_df_l = duckdb.sql(extra_df_l_sql).to_df()
    extra_df_r_sql = f"""
    SELECT concat_main.* from (SELECT concat_ws('▼',{','.join(str(item) for item in keys)}) as ConcatenatedKeys,* from '{datafile_r}') AS concat_main JOIN '{outputfile_r}' extra_r on concat_main.ConcatenatedKeys = extra_r.ConcatenatedKeys limit {compare_row_limit}
    """
    extra_df_r = duckdb.sql(extra_df_r_sql).to_df()
    return extra_df_l, extra_df_r
  
    
def get_common_rows_using_keys_duckdb(filename_l, filename_r, common_file_name, datafile_l, datafile_r, keys, common_col_keys):
    common_file_name=str(common_file_name)
    query = f"""
    SELECT count(1)
    FROM (
    SELECT ConcatenatedKeys,HashValue FROM '{filename_l}'
    INTERSECT
    SELECT ConcatenatedKeys,HashValue FROM '{filename_r}'
    ) AS common_tbl;
    """
    count = duckdb.sql(query).to_df().values[0][0]
    ic(count)
    query = f"""
    SELECT *
    FROM (
    SELECT ConcatenatedKeys,HashValue FROM '{filename_l}'
    INTERSECT
    SELECT ConcatenatedKeys,HashValue FROM '{filename_r}'
    ) AS common_tbl limit 10;
    """
    duckdb.sql(query).to_parquet(common_file_name,compression='gzip')
    common_rows_sql = f"""
    SELECT left_concat_main.*, right_concat_main.* from (SELECT concat_ws('▼',{','.join(str(item) for item in keys)}) as ConcatenatedKeys_x, {','.join(str(item)+ ' AS '+str(item)+'_x' for item in common_col_keys)} from '{datafile_l}') AS left_concat_main, (SELECT concat_ws('▼',{','.join(str(item) for item in keys)}) as ConcatenatedKeys_y, {','.join(str(item)+ ' AS '+str(item)+'_y' for item in common_col_keys)} from '{datafile_r}') AS right_concat_main , '{common_file_name}' AS common_file where left_concat_main.ConcatenatedKeys_x = common_file.ConcatenatedKeys and left_concat_main.ConcatenatedKeys_x = right_concat_main.ConcatenatedKeys_y limit 10
    """
    common_rows = duckdb.sql(common_rows_sql).to_df()
    ic(common_rows)
    return count, common_rows


def get_common_rows(df_left, df_right, keys):
    """
    Get the common rows between two dataframes based on specified keys.

    Parameters:
    df_left (pandas.DataFrame): First dataframe.
    df_right (pandas.DataFrame): Second dataframe.
    keys (list): List of column names to use as keys for comparison.

    Returns:
    pandas.DataFrame: A dataframe containing the common rows between df_left and df_right.
    """
    # Merge the dataframes on the specified keys
    merged_df = pd.merge(df_left, df_right, on=keys)

    return merged_df

def get_mismatched_fields(df1, df2, mismatched_df):
    """
    Get the mismatched fields for each mismatched record between two dataframes.

    Parameters:
    df1 (pandas.DataFrame): First original dataframe.
    df2 (pandas.DataFrame): Second original dataframe.
    mismatched_df (pandas.DataFrame): Dataframe containing the mismatched rows.

    Returns:
    dict: A dictionary where the keys are the index of the mismatched records and the values are lists of mismatched fields.
    """
    mismatched_fields = {}

    for index, row in mismatched_df.iterrows():
        mismatched_fields[index] = []

        for column in mismatched_df.columns:
            if column in ["ConcatenatedKeys", "HashValue"]:
                if row[column] != df1.loc[index, column]:
                    mismatched_fields[index].append(column)

    return mismatched_fields


def compare_mismatched_rows_old(
    df1: pd.DataFrame, df2: pd.DataFrame, mismatched_df: pd.DataFrame
) -> benedict[str, benedict]:
    """
    Compare mismatched rows in two dataframes and return a dictionary with differing columns.

    Parameters:
    df1 (pandas.DataFrame): First dataframe.
    df2 (pandas.DataFrame): Second dataframe.
    mismatched_df (pandas.DataFrame): Dataframe with mismatched rows.

    Returns:
    benedict: A dictionary with column names as keys and dataframes as values.
    """
    result_dict = benedict()

    for index, row in mismatched_df.iterrows():
        keys = row["ConcatenatedKeys"]
        df1_row = df1[df1["ConcatenatedKeys"] == keys]
        df2_row = df2[df2["ConcatenatedKeys"] == keys]

        for column in df1.columns:
            if (
                column not in ["ConcatenatedKeys", "HashValue"]
                and df1_row[column].values[0] != df2_row[column].values[0]
            ):
                if column not in result_dict:
                    result_dict[column] = []
                result_dict[column].append(
                    {
                        "df1_value": df1_row[column].values[0],
                        "df2_value": df2_row[column].values[0],
                        "ConcatenatedKeys": keys,
                    }
                )

    return result_dict

def compare_mismatched_rows(
    df_left: pd.DataFrame, df_right: pd.DataFrame, mismatched_df: pd.DataFrame, df_left_extra : pd.DataFrame,df_right_extra : pd.DataFrame
) -> benedict[str, benedict]:
    """
    Compare mismatched rows in two dataframes and return a dictionary with differing columns.

    Parameters:
    df_left (pandas.DataFrame): First dataframe.
    df_right (pandas.DataFrame): Second dataframe.
    mismatched_df (pandas.DataFrame): Dataframe with mismatched rows.

    Returns:
    benedict: A dictionary with column names as keys and dataframes as values.
    """
    result_dict = benedict()

    for index, row in mismatched_df.iterrows():
        keys = row["ConcatenatedKeys"]
        df_left_row = df_left[df_left["ConcatenatedKeys"] == keys]
        df_right_row = df_right[df_right["ConcatenatedKeys"] == keys]

        for column in df_left.columns:
            if (
                column not in ["ConcatenatedKeys", "HashValue"]
                and df_left_row[column].values[0] != df_right_row[column].values[0]
            ):
                if column not in result_dict:
                    result_dict[column] = []
                result_dict[column].append(
                    {
                        "df_left_value": df_left_row[column].values[0],
                        "df_right_value": df_right_row[column].values[0],
                        "ConcatenatedKeys": keys,
                    }
                )
    for index, row in df_left_extra.iterrows():
        keys = row["ConcatenatedKeys"]
        df_left_row = df_left[df_left["ConcatenatedKeys"] == keys]
        for column in df_left.columns:
            if (
                column not in ["ConcatenatedKeys", "HashValue"]
            ):
                if column not in result_dict:
                    result_dict[column] = []
                result_dict[column].append(
                    {
                        "df_left_value": df_left_row[column].values[0],
                        "df_right_value": "Does_Not_Exists",
                        "ConcatenatedKeys": keys,
                    }
                )
    for index, row in df_right_extra.iterrows():
        keys = row["ConcatenatedKeys"]
        df_right_row = df_right[df_right["ConcatenatedKeys"] == keys]
        for column in df_right.columns:
            if (
                column not in ["ConcatenatedKeys", "HashValue"]
            ):
                if column not in result_dict:
                    result_dict[column] = []
                result_dict[column].append(
                    {
                        "df_left_value": "Does_Not_Exists",
                        "df_right_value": df_right_row[column].values[0],
                        "ConcatenatedKeys": keys,
                    }
                )

    return result_dict

def find_mismatched_rows(df_left, df_right):
    """
    Compare two dataframes based on the 'ConcatenatedKeys' column and return rows where 'HashValue' is not matching.

    Parameters:
    df_left (pandas.DataFrame): First dataframe.
    df_right (pandas.DataFrame): Second dataframe.

    Returns:
    pandas.DataFrame: A dataframe containing the rows where 'HashValue' is not matching based on the 'ConcatenatedKeys' column.
    """
    merged_df = pd.merge(df_left, df_right, on="ConcatenatedKeys", suffixes=("_df_left", "_df_right"))
    ic(merged_df)
    mismatched_df = merged_df[merged_df["HashValue_df_left"] != merged_df["HashValue_df_right"]]
    ic(mismatched_df)
    
    # Convert pandas dataframes to DuckDB tables
    con = duckdb.connect()
    con.register("df_left", df_left)
    con.register("df_right", df_right)

    # Perform the merge using DuckDB
    merged_df = con.execute("""
        SELECT *
        FROM df_left
        JOIN df_right ON df_left.ConcatenatedKeys = df_right.ConcatenatedKeys
    """).fetchdf()

    # Filter the mismatched rows using DuckDB
    mismatched_df = con.execute("""
        SELECT *
        FROM merged_df
        WHERE merged_df.HashValue_df_left != merged_df.HashValue_df_right
    """).fetchdf()

    return mismatched_df


def find_mismatched_rows_duckdb(filename_l, filename_r, diff_file_name, datafile_l, datafile_r, keys, common_col_keys):
    diff_file_name=str(diff_file_name)
    query = f"""
    SELECT count(1)
    FROM (
    SELECT a.ConcatenatedKeys,a.HashValue as 'HashValue_df_left',b.HashValue as 'HashValue_df_right'
    FROM '{filename_l}' as a JOIN '{filename_r}' as b ON a.ConcatenatedKeys = b.ConcatenatedKeys
    WHERE
    a.HashValue != b.HashValue
    ) AS common_tbl;
    """
    count = duckdb.sql(query).to_df().values[0][0]
    ic(count)
    query = f"""
    SELECT *
    FROM (
    SELECT a.ConcatenatedKeys,a.HashValue as 'HashValue_df_left',b.HashValue as 'HashValue_df_right'
    FROM '{filename_l}' as a JOIN '{filename_r}' as b ON a.ConcatenatedKeys = b.ConcatenatedKeys
    WHERE
    a.HashValue != b.HashValue
    ) AS DiffTbl limit 10;
    """
    duckdb.sql(query).to_parquet(diff_file_name,compression='gzip')
    diff_rows_sql = f"""
    SELECT left_concat_main.*, right_concat_main.* 
    from 
    (SELECT concat_ws('▼',{','.join(str(item) for item in keys)}) as ConcatenatedKeys_x, {','.join(str(item)+ ' AS '+str(item)+'_x' for item in common_col_keys)} from '{datafile_l}') AS left_concat_main, 
    (SELECT concat_ws('▼',{','.join(str(item) for item in keys)}) as ConcatenatedKeys_y, {','.join(str(item)+ ' AS '+str(item)+'_y' for item in common_col_keys)} from '{datafile_r}') AS right_concat_main, 
    '{diff_file_name}' AS diff_file
    where left_concat_main.ConcatenatedKeys_x = diff_file.ConcatenatedKeys and left_concat_main.ConcatenatedKeys_x = right_concat_main.ConcatenatedKeys_y limit 10
    """
    common_rows = duckdb.sql(diff_rows_sql).to_df()
    ic(common_rows)
    return count, common_rows


    # Perform the merge using DuckDB
    merged_df = duckdb.sql("""
        SELECT *
        FROM df_left
        JOIN df_right ON df_left.ConcatenatedKeys = df_right.ConcatenatedKeys
    """).to_df()

    # Filter the mismatched rows using DuckDB
    mismatched_df = duckdb.sql("""
        SELECT *
        FROM merged_df
        WHERE merged_df.HashValue_df_left != merged_df.HashValue_df_right
    """).fetchdf()



    diff_file_name=str(diff_file_name)
    query = f"""
    SELECT count(1)
    FROM (
    SELECT ConcatenatedKeys,HashValue FROM '{filename_l}'
    INTERSECT
    SELECT ConcatenatedKeys,HashValue FROM '{filename_r}'
    ) AS common_tbl;
    """
    count = duckdb.sql(query).to_df().values[0][0]
    ic(count)
    query = f"""
    SELECT *
    FROM (
    SELECT ConcatenatedKeys,HashValue FROM '{filename_l}'
    INTERSECT
    SELECT ConcatenatedKeys,HashValue FROM '{filename_r}'
    ) AS common_tbl limit 10;
    """
    duckdb.sql(query).to_parquet(diff_file_name,compression='gzip')
    diff_rows_sql = f"""
    SELECT left_concat_main.*, right_concat_main.* from (SELECT concat_ws('▼',{','.join(str(item) for item in keys)}) as ConcatenatedKeys_x, {','.join(str(item)+ ' AS '+str(item)+'_x' for item in keys)} from '{datafile_l}') AS left_concat_main, (SELECT concat_ws('▼',{','.join(str(item) for item in keys)}) as ConcatenatedKeys_y, {','.join(str(item)+ ' AS '+str(item)+'_y' for item in keys)} from '{datafile_r}') AS right_concat_main , '{diff_file_name}' AS common_file where left_concat_main.ConcatenatedKeys_x = common_file.ConcatenatedKeys and left_concat_main.ConcatenatedKeys_x = right_concat_main.ConcatenatedKeys_y limit 10
    """
    common_rows = duckdb.sql(diff_rows_sql).to_df()
    ic(common_rows)
    return count, common_rows

def highlight_positive_cells(excel_filename, worksheet_name, column_name):
    """
    Highlights cells with positive values in the specified column of an Excel worksheet.

    Args:
        excel_filename (str): Path to the Excel file.
        worksheet_name (str): Name of the worksheet.
        column_name (str): Name of the column to check for positive values.

    Returns:
        None
    """
    try:
        # Read the Excel file
        df = pd.read_excel(excel_filename, sheet_name=worksheet_name)

        # Get the column index based on the column name
        column_index = df.columns.get_loc(column_name)

        # Create a workbook object
        wb = openpyxl.load_workbook(excel_filename)

        # Select the specified worksheet
        ws = wb[worksheet_name]

        # Define the red fill style
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

        # Iterate through rows and highlight positive cells
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=column_index + 1, max_col=column_index + 1):
            cell_value = row[0].value
            if cell_value is not None and cell_value > 0:
                row[0].fill = red_fill

        # Save the modified workbook
        wb.save(excel_filename)
        print(f"Positive values in column '{column_name}' highlighted in red.")

    except Exception as e:
        print(f"Error: {str(e)}")


def highlight_cells_with_value(excel_filename, target_value):
    """
    Highlights cells in an Excel file that match the specified value.

    Args:
        excel_filename (str): Path to the Excel file.
        target_value (str): The value to search for in the cells.

    Returns:
        None
    """
    try:
        # Create a workbook object
        wb = openpyxl.load_workbook(excel_filename)

        # Iterate through all sheets in the workbook
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Define the red fill style
            red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

            # Iterate through all cells in the sheet
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value == target_value:
                        cell.fill = red_fill

        # Save the modified workbook
        wb.save(excel_filename)
        print(f"Cells with value '{target_value}' highlighted in red.")

    except Exception as e:
        print(f"Error: {str(e)}")


def make_header_background_grey(excel_filename):
    """
    Modifies the background color of the header row in an Excel file to grey (if header cell has text).

    Args:
        excel_filename (str): Path to the Excel file.

    Returns:
        None
    """
    try:
        # Create a workbook object
        wb = openpyxl.load_workbook(excel_filename)

        # Iterate through all sheets in the workbook
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Define the grey fill style
            grey_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

            # Iterate through header cells (first row)
            for cell in ws[1]:
                if cell.value:
                    cell.fill = grey_fill

        # Save the modified workbook
        wb.save(excel_filename)
        print("Header background color set to grey for cells with text.")

    except Exception as e:
        print(f"Error: {str(e)}")

def get_last_line(file_path):
    """
    Get the last line of a file using the tail command.

    Parameters:
    file_path (str): Path to the file.

    Returns:
    str: The last line of the file.
    """
    tail_command = f"tail -n 1 {file_path}"
    result = subprocess.run(tail_command, shell=True, capture_output=True, text=True)
    last_line = result.stdout.strip()
    return last_line



def string_to_list(string: str) -> list[str]:
    if not string or len(string) == 0:
        return []
    if ',' not in string:
        return [string]
    return string.split(',')






def get_filename_without_extension(fullpath):
    """
    Get the filename without extension from a full path.

    Parameters:
    fullpath (str): The full path of the file.

    Returns:
    str: The filename without extension.
    """
    filename = os.path.basename(fullpath)
    filename_without_extension = os.path.splitext(filename)[0]
    return filename_without_extension



def create_csv_file(colmapping, filename1,sep):
    df = create_large_df(50, list(colmapping['filecolname']), list(colmapping['type']), list(colmapping['length']))
    header1 = 'filename\n'
    header2 = ','.join(df.columns)
    # Remove filename1 if exists
    if os.path.exists(filename1):
        os.remove(filename1)
        
    start_time = tx.perf_counter()
    ic(sep)

    df.to_csv(filename1, index=False, sep=sep, header=list(df.columns), quotechar='"', quoting=1, escapechar='\\')

    end_time = tx.perf_counter()
    execution_time = end_time - start_time
    print(f"Execution time: {execution_time} seconds")

    # Add trailer with record count
    record_count = len(df)
    trailer = f'RECORDCOUNT={record_count}\n'

    with open(filename1, 'a') as file:
        file.write(trailer)
    # with open(filename1, 'r+') as f:
    #     content = f.read()
    #     f.seek(0, 0)
    #     f.write('new first line\n' + content)
    temp_filename = 'temp.txt'   
    if os.path.exists(temp_filename):
        os.remove(temp_filename)
    with open(filename1, 'r') as original_file, open(temp_filename, 'w') as temp_file:
        temp_file.write(header1)
        shutil.copyfileobj(original_file,temp_file)
    ic(filename1)
    os.remove(filename1)
    ic(temp_filename)
    shutil.move(temp_filename, filename1)
    shutil.copy(filename1, f'{filename1}.bak2')
    ic(filename1)

    script_path = '/home/deck/devlopment/demo/fixedlenghtconv.sh'
    subprocess.run([script_path, filename1])
    if os.path.exists(temp_filename):
        os.remove(temp_filename)



def create_large_df(num_rows, columns=['A', 'B'], column_types=['int', 'int'],column_lengnths=[10, 10]):
    fake = Faker()
    # Create a dataframe with random values
    df = pd.DataFrame()
    start = pendulum.datetime(1970, 1, 1)
    end = pendulum.now()
    delta = end - start
    i = 0
    for col, col_type, col_length in zip(columns, column_types, column_lengnths):
        ic(i)
        if col_type == 'int':
            if i == 0:
                df[col] = [str(i).zfill(col_length) for i in range(1, num_rows+1)]
            else:
                df[col] = [str(random.randint(1, 1000000)).zfill(col_length)[:col_length] for _ in range(num_rows)]
        elif col_type == 'float':
            df[col] = [str(fake.pydecimal(left_digits=15, right_digits=2, positive=True)).zfill(col_length)[-col_length:] for _ in range(num_rows)]
        elif col_type == 'str':
            df[col] = [fake.name().ljust(col_length)[:col_length] for _ in range(num_rows)]
        elif col_type == 'date':
            df[col] = [start.add(seconds=random.randint(0, delta.in_seconds())).to_date_string() for _ in range(num_rows)]
        elif col_type == 'datetime':
            df[col] = [start.add(seconds=random.randint(0, delta.in_seconds())).to_datetime_string() for _ in range(num_rows)]
        i += 1
    return df


def delete_file_if_exists(file_path):
    """
    Delete the file if it exists.

    Parameters:
    file_path (str): Path to the file.

    Returns:
    None
    """
    if os.path.exists(file_path):
        os.remove(file_path)
        print(f"File '{file_path}' deleted.")
    else:
        print(f"File '{file_path}' does not exist.")


def get_paraquete_rowcount(file_path):
    duckdb.query(f"SELECT count(1) as cnt FROM '{file_path}'").to_df().values[0][0]



def main():
    config_path = '/home/deck/Documents/mappings/test1_F2F.xlsx'
    colmapping, meta_l, meta_r, filename_l, filename_r = read_config(config_path,'abc.txt','pqr.txt')
    list_of_columns = list(colmapping['filecolname'])
    uniq_col_list = string_to_list(colmapping['uniq_key'][0])
    ic(uniq_col_list)
    lenghtlist = list(colmapping['length'])

    run_id = uuid.UUID(int=uuid.getnode()).hex[-12:]
    run_id = 6 


    ic(list_of_columns)

    
    try:
        fixed_l = meta_l.loc[meta_l['MKey'] == 'FIXLENGTH', 'MValue'].values[0]
    except:
        try:
            fixed_l = fixed_l = meta_l.loc[meta_l['MKey'] == 'FIXLENGHT', 'MValue'].values[0]
        except:
            fixed_l = 'N'



    if fixed_l == 'N':
        delmt_l = ''
        delmt_l = meta_l.loc[meta_l['MKey'] == 'DELIMETER', 'MValue'].values[0]
        if delmt_l == '\|':
            delmt_l = '|'
        if delmt_l == '\\|':
            delmt_l = '|'
        try:
            quote_char_l = meta_l.loc[meta_l['MKey'] == 'QUOTEWRAPPER', 'MValue'].values[0]
        except:
            quote_char_l = ''
        try:
            escape_char_l = meta_l.loc[meta_l['MKey'] == 'ESCAPECHAR', 'MValue'].values[0]
        except:
            escape_char_l = ''  
    else:
        delmt_l = ''
        quote_char_l = ''
        escape_char_l = ''

    try:
        headrows_l = meta_l.loc[meta_l['MKey'] == 'HEADER_ROWS', 'MValue'].values[0]
    except:
        headrows_l = '0'

    try:
        tailrows_l = meta_l.loc[meta_l['MKey'] == 'TRAILER_ROWS', 'MValue'].values[0]
    except:
        tailrows_l = '0' 


    try:
        trailercntstr_l = meta_l.loc[meta_l['MKey'] == 'TRAILERCNTSTR', 'MValue'].values[0]
    except:
        trailercntstr_l = ''



    try:
        fixed_r = meta_r.loc[meta_r['MKey'] == 'FIXLENGTH', 'MValue'].values[0]
    except:
        try:
            fixed_r = meta_r.loc[meta_r['MKey'] == 'FIXLENGHT', 'MValue'].values[0]
        except:
            fixed_r = 'N'

    if fixed_r == 'N':
        delmt_r = ''
        delmt_r = meta_r.loc[meta_r['MKey'] == 'DELIMETER', 'MValue'].values[0]
        if delmt_r == '\|':
            delmt_r = '|'
        if delmt_r == '\\|':
            delmt_r = '|'
        try:
            quote_char_r = meta_r.loc[meta_r['MKey'] == 'QUOTEWRAPPER', 'MValue'].values[0]
        except:
            quote_char_r = ''
        try:
            escape_char_r = meta_r.loc[meta_r['MKey'] == 'ESCAPECHAR', 'MValue'].values[0]
        except:
            escape_char_r = ''  
    else:
        delmt_r = ''
        quote_char_r = ''
        escape_char_r = ''

    try:
        headrows_r = meta_r.loc[meta_r['MKey'] == 'HEADER_ROWS', 'MValue'].values[0]
    except:
        headrows_r = '0'

    try:
        tailrows_r = meta_r.loc[meta_r['MKey'] == 'TRAILER_ROWS', 'MValue'].values[0]
    except:
        tailrows_r = '0' 


    try:
        trailercntstr_r = meta_r.loc[meta_r['MKey'] == 'TRAILERCNTSTR', 'MValue'].values[0]
    except:
        trailercntstr_r = ''



    script_path = '/home/deck/devlopment/demo/cleanheadandtail.sh'
    param1 = f'{filename_l}'
    param2 = str(headrows_l)
    param3 = str(tailrows_l)
    subprocess.run([script_path, param1, param2, param3])
    if fixed_l == 'N':
        df_l = pd.read_csv(f'{filename_l}.gz',engine='pyarrow', dtype_backend='pyarrow', delimiter=delmt_l, names=list_of_columns, skiprows=0, header=None, encoding='utf-8' ,escapechar=escape_char_l, compression='gzip')
    else:
        df_l = pd.read_fwf(f'{filename_l}.gz', widths=lenghtlist, names=list_of_columns,engine='pyarrow', dtype_backend='pyarrow', compression='gzip')
    try:
        last_line_l=get_last_line(filename_l.replace('.left',''))
        last_line_l=last_line_l.replace(trailercntstr_l,'').strip().replace('=','')
        last_line_i_l = int(last_line_l)
    except:
        last_line_i_l = 0

    delete_file_if_exists(f'{filename_l}_{run_id}.parquet')
    df_l.to_parquet(f'{filename_l}_{run_id}.parquet', engine='pyarrow', compression='gzip')

    script_path = '/home/deck/devlopment/demo/cleanheadandtail.sh'
    param1 = f'{filename_r}'
    param2 = str(headrows_r)
    param3 = str(tailrows_r)
    subprocess.run([script_path, param1, param2, param3])
    if fixed_r == 'N':
        df_r = pd.read_csv(f'{filename_r}.gz',engine='pyarrow', dtype_backend='pyarrow', delimiter=delmt_r, names=list_of_columns, skiprows=0, header=None, encoding='utf-8' ,escapechar=escape_char_r, compression='gzip')
    else:
        df_r = pd.read_fwf(f'{filename_r}.gz', widths=lenghtlist, names=list_of_columns,engine='pyarrow', dtype_backend='pyarrow', compression='gzip')
    try:
        last_line_r=get_last_line(filename_r.replace('.right',''))
        last_line_r=last_line_r.replace(trailercntstr_r,'').strip().replace('=','')
        last_line_i_r = int(last_line_r)
    except:
        last_line_i_r = 0

    delete_file_if_exists(f'{filename_r}_{run_id}.parquet')
    df_r.to_parquet(f'{filename_r}_{run_id}.parquet', engine='pyarrow', compression='gzip')

    

    ic(last_line_i_l)
    ic(last_line_i_r)

    #hash_df_l = create_hash_dataframe(df_l.copy(), uniq_col_list)
    #hash_df_r = create_hash_dataframe(df_r.copy(), uniq_col_list)
    
    delete_file_if_exists(f'{filename_l}_{run_id}_hash.parquet')
    create_hash_dataframe_duckdb(f'{filename_l}_{run_id}.parquet',uniq_col_list,list_of_columns,f'{filename_l}_{run_id}_hash.parquet')
    delete_file_if_exists(f'{filename_r}_{run_id}_hash.parquet')
    create_hash_dataframe_duckdb(f'{filename_r}_{run_id}.parquet',uniq_col_list,list_of_columns,f'{filename_r}_{run_id}_hash.parquet')
    get_extra_rows_using_keys_duckdb(f'{filename_l}_{run_id}_hash.parquet',f'{filename_r}_{run_id}_hash.parquet',f'{filename_l}_{run_id}_extra.parquet',f'{filename_r}_{run_id}_extra.parquet',f'{filename_l}_{run_id}.parquet',f'{filename_r}_{run_id}.parquet',uniq_col_list)
    common_count,common_rows = get_common_rows_using_keys_duckdb(f'{filename_l}_{run_id}_hash.parquet',f'{filename_r}_{run_id}_hash.parquet',str(Path(Path(f'{filename_l}_{run_id}_hash.parquet').parent,f'{run_id}_common.parquet')), f'{filename_l}_{run_id}.parquet',f'{filename_r}_{run_id}.parquet',uniq_col_list,list_of_columns)
    find_mismatched_rows_duckdb(f'{filename_l}_{run_id}_hash.parquet',f'{filename_r}_{run_id}_hash.parquet',str(Path(Path(f'{filename_l}_{run_id}_hash.parquet').parent,f'{run_id}_common.parquet')), f'{filename_l}_{run_id}.parquet',f'{filename_r}_{run_id}.parquet',uniq_col_list,list_of_columns)

    # extra_df_l, extra_df_r = get_extra_rows_using_keys(
    #     df_l.copy(), df_r.copy(),uniq_col_list
    # )
    # extra_df_l, extra_df_r = get_extra_rows_using_hash(hash_df_l, hash_df_r)
    
    # ic(extra_df_l[:10])
    # ic(extra_df_r[:10])
    # common_df = get_common_rows(hash_df_l, hash_df_r, ["ConcatenatedKeys", "HashValue"])
    # print(common_df)
    
    mismatched_df = find_mismatched_rows(
        hash_df_l[["ConcatenatedKeys", "HashValue"]].copy(),
        hash_df_r[["ConcatenatedKeys", "HashValue"]].copy(),
    )
    ic(mismatched_df)

    #To DO this needs to be fulldf: there needs to be a way for unselecting batch, only one key column not working
    mismatched_dict = compare_mismatched_rows(hash_df_l, hash_df_r, mismatched_df,extra_df_l,extra_df_r)
    ic(mismatched_dict)

    passed_col_df = pd.DataFrame(columns=['column_name', 'ConcatenatedKeys', 'value_df_left', 'value_df_right'])
    ic(passed_col_df)
    passed_benedict_list : list[benedict] = []
    for column in df_l.columns:
        #ic(column)
        for index, common_row in common_df.head(10).iterrows():
                    # New row as a DataFrame
            if column in ['ConcatenatedKeys','HashValue']:
                continue
            item =  benedict()
            item.column_name = column
            item.ConcatenatedKeys = common_row.ConcatenatedKeys
            item.value_df_left = common_row[f'{column}_x']
            item.value_df_right = common_row[f'{column}_y']
            passed_benedict_list.append(item)
            

    ic(passed_benedict_list)

    passed_col_df = pd.DataFrame(passed_benedict_list)
    path = f'/home/deck/Downloads/'
    report_path = Path(path,f'{get_filename_without_extension(config_path)}_report_{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx')
    with pd.ExcelWriter(report_path, engine="xlsxwriter") as writer:
        # Create a new dataframe with the column names from df_left
        summary_df = pd.DataFrame({"ColumnName": df_l.columns})

        # Write the dataframe to the Excel sheet
        summary_df["MatchedRowsCountLeft"] = summary_df["ColumnName"].apply(
            lambda col: len(hash_df_l)
            - len(extra_df_l)
            - ((len(mismatched_dict[col]) if col in mismatched_dict else 0) - (len(extra_df_l)+len(extra_df_r)))
        )

        summary_df["MatchedRowsCountRight"] = summary_df["ColumnName"].apply(
            lambda col: len(hash_df_r)
            - len(extra_df_r)
            - ((len(mismatched_dict[col]) if col in mismatched_dict else 0) - (len(extra_df_l)+len(extra_df_r)))
        )


        summary_df["MisMatchedRowsCount"] = summary_df["ColumnName"].apply(
            lambda col: (len(mismatched_dict[col]) if col in mismatched_dict else 0)
        )

        summary_df["ExtraInLeft"] = summary_df["ColumnName"].apply(
            lambda col: len(extra_df_l)
        )
        summary_df["ExtraInRight"] = summary_df["ColumnName"].apply(
            lambda col: len(extra_df_r)
        )

        summary_df["TotalRowsInLeft"] = len(hash_df_l)
        summary_df["TotalRowsInRight"] = len(hash_df_r)

        
        summary_df = summary_df.drop([0, 1])
        meta_df = pd.DataFrame(columns=['ConfigName', 'ConfigValue'])
        meta_df.loc[0] = ['LeftFileName', filename_l.replace('.left','')]
        meta_df.loc[1] = ['LeftTableName', filename_r.replace('.right','')]



        meta_df.to_excel(writer, sheet_name="MetaInfo", index=False)

        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        passed_col_df.to_excel(writer, sheet_name="PassRecordSummary", index=False)

        
        
        for column_name, differ_items in mismatched_dict.items():
            mismatched_df = pd.DataFrame(columns=['ConcatenatedKeys', 'value_df_left', 'value_df_right'])
            mismatched_bendict_list : list[benedict] = []
            for dit in differ_items:
                item =  benedict()
                item.ConcatenatedKeys = dit.ConcatenatedKeys
                item.value_df_left = dit.df_left_value
                item.value_df_right = dit.df_right_value
                mismatched_bendict_list.append(item)
            mismatched_df = pd.DataFrame(mismatched_bendict_list)

            ic(extra_df_l)
            mismatched_df.to_excel(writer, sheet_name=f"{column_name}_miss", index=False)
    highlight_positive_cells(report_path,'Summary','MisMatchedRowsCount')
    highlight_positive_cells(report_path,'Summary','ExtraInLeft')
    highlight_positive_cells(report_path,'Summary','ExtraInRight')
    highlight_cells_with_value(report_path,'Does_Not_Exists')
    make_header_background_grey(report_path)
    ic(report_path)



if __name__ == "__main__":
    main()