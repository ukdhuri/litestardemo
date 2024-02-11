import json
import queue
import subprocess
from typing import Optional
from benedict import benedict
from faker import Faker
from loguru import logger
from omegaconf import DictConfig
import pandas as pd
from sqlalchemy import Engine
from sqlmodel import SQLModel, create_engine, select
from datetime import datetime,date,time,timedelta
from litestar.datastructures import State
from hydra import compose, initialize_config_dir
from models import local
import os
# from sqlalchemy.ext.asyncio  import AsyncSession
from sqlmodel.ext.asyncio.session import AsyncSession
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine
from models import remote
from sqlalchemy.exc import IntegrityError, NoResultFound
from litestar.exceptions import ClientException, NotFoundException,InternalServerException
from litestar.status_codes import HTTP_409_CONFLICT
from typing import TypeVar
from typing import List
from icecream import ic
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

initialize_config_dir(version_base=None, config_dir=f"{os.getcwd()}/config", job_name="demo")
cfg = compose(config_name="config")
lstate=State({"appcfg": cfg})
T = TypeVar('T')

def decrpyt_password(cybeark_string):
    p = subprocess.Popen(cybeark_string, shell=True, stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
    output, _ = p.communicate()
    output_str = output.decode('utf-8')
    if "ERROR" in output_str.upper():
        print("Error")
        exit(1)
    return output_str.strip()

def get_engine(cfg: DictConfig, key: str)  -> Engine:
    connection_string = cfg[key].vendor.connection_string
    if key != "local":
        password=decrpyt_password(cfg[key].password_command)
        connection_string = connection_string.replace("REPLACEME", password)
    engine = create_engine(connection_string, echo=False)
    return engine

def get_connection_strnig(key: str)  -> str:
    connection_string = cfg[key].vendor.connection_string
    if key != "local":
        password=decrpyt_password(cfg[key].password_command)
        connection_string = connection_string.replace("REPLACEME", password)
    return connection_string

def generate_dates(start_date, end_date):
    delta = end_date - start_date
    for i in range(delta.days + 1):
        date = start_date + timedelta(days=i)
        yield date


def reset_queue(q):
    q = queue.Queue()
    for i in range(10000):
        q.put(i+1)
    return q

async def on_startup():
    logger.debug("App started!")
    #logger.debug("That's it, beautiful and simple Started!")

sessionmaker = async_sessionmaker(expire_on_commit=False)
async def get_todo_list(done: Optional[bool], session: AsyncSession) -> list[remote.User]:
    query = select(remote.User)
    print(query)
    async with sessionmaker(bind=session.owner.state.engine) as session1:
        try:
            async with session1.begin():
                result = await session1.exec(query)
        except IntegrityError as exc:
            raise ClientException(
                status_code=HTTP_409_CONFLICT,
                detail=str(exc),
            ) from exc
    return result.scalars().all()


async def get_todo_listX(done: Optional[bool], session: AsyncSession) -> list[remote.User]:
    query = select(remote.User).order_by(remote.User.id.asc()).offset(0).limit(5)
    if done is not None:
        query = query.where(remote.User > 1)

    result = await session.exec(query)
    return result.scalars().all()

async def get_all_users(session: AsyncSession) -> list[remote.User]:
   return session.exec(select(remote.User)).scalars().all()



async def getemoji() -> str:
    fake = Faker()
    return fake.emoji()


def list_to_string(lst : List[str]) -> str:
    if not lst or len(lst) == 0:
        return ''
    return ','.join(str(item) for item in lst)

def string_to_list(string: str) -> List[str]:
    if not string or len(string) == 0:
        return []
    if ',' not in string:
        return [string]
    return string.split(',')

def list_to_json(lst : List[str]) -> str:
    if not lst or len(lst) == 0:
        return '[]'
    return json.dumps(lst)

def json_to_list(string: str) -> List[str]:
    if not string or len(string) == 0:
        return []
    return json.loads(string)


def check_three_vars(a, b, c):
    if (a == '' and b == '' and (c == '' or c == 'Select Date Format') ) or (a != '' and b != '' and (c != 'Select Date Format' or c!= ''))  :
        return True
    else:
        return False
    
def get_extra_rows_using_hash(df_left, df_right):
    """
    Compare two dataframes and return extra rows based on the 'ConcatenatedKeys' column.

    Parameters:
    df_left (pandas.DataFrame): First dataframe.
    df_right (pandas.DataFrame): Second dataframe.

    Returns:
    tuple: A tuple containing two dataframes with extra rows from df_left and df_right.
    """
    df_left_extra = df_left[~df_left["ConcatenatedKeys"].isin(df_right["ConcatenatedKeys"])]
    df_right_extra = df_right[~df_right["ConcatenatedKeys"].isin(df_left["ConcatenatedKeys"])]

    return df_left_extra, df_right_extra

def get_common_rows(df_left, df_right, keys):
    """
    Get the common rows between two dataframes based on specified keys.

    Parameters:
    df_left (pandas.DataFrame): First dataframe.
    df_right (pandas.DataFrame): Second dataframe.
    keys (list): List of column names to use as keys for comparison.

    Returns:
    pandas.DataFrame: A dataframe containing the common rows between df_left and df_right.
    """
    # Merge the dataframes on the specified keys
    merged_df = pd.merge(df_left, df_right, on=keys)

    return merged_df

def compare_mismatched_rows(
    df_left: pd.DataFrame, df_right: pd.DataFrame, mismatched_df: pd.DataFrame, df_left_extra : pd.DataFrame,df_right_extra : pd.DataFrame
) -> benedict[str, benedict]:
    """
    Compare mismatched rows in two dataframes and return a dictionary with differing columns.

    Parameters:
    df_left (pandas.DataFrame): First dataframe.
    df_right (pandas.DataFrame): Second dataframe.
    mismatched_df (pandas.DataFrame): Dataframe with mismatched rows.

    Returns:
    benedict: A dictionary with column names as keys and dataframes as values.
    """
    result_dict = benedict()

    for index, row in mismatched_df.iterrows():
        keys = row["ConcatenatedKeys"]
        df_left_row = df_left[df_left["ConcatenatedKeys"] == keys]
        df_right_row = df_right[df_right["ConcatenatedKeys"] == keys]

        for column in df_left.columns:
            if (
                column not in ["ConcatenatedKeys", "HashValue"]
                and df_left_row[column].values[0] != df_right_row[column].values[0]
            ):
                if column not in result_dict:
                    result_dict[column] = []
                result_dict[column].append(
                    {
                        "df_left_value": df_left_row[column].values[0],
                        "df_right_value": df_right_row[column].values[0],
                        "ConcatenatedKeys": keys,
                    }
                )
    for index, row in df_left_extra.iterrows():
        keys = row["ConcatenatedKeys"]
        df_left_row = df_left[df_left["ConcatenatedKeys"] == keys]
        for column in df_left.columns:
            if (
                column not in ["ConcatenatedKeys", "HashValue"]
            ):
                if column not in result_dict:
                    result_dict[column] = []
                result_dict[column].append(
                    {
                        "df_left_value": df_left_row[column].values[0],
                        "df_right_value": "Does_Not_Exists",
                        "ConcatenatedKeys": keys,
                    }
                )
    for index, row in df_right_extra.iterrows():
        keys = row["ConcatenatedKeys"]
        df_right_row = df_right[df_right["ConcatenatedKeys"] == keys]
        for column in df_right.columns:
            if (
                column not in ["ConcatenatedKeys", "HashValue"]
            ):
                if column not in result_dict:
                    result_dict[column] = []
                result_dict[column].append(
                    {
                        "df_left_value": "Does_Not_Exists",
                        "df_right_value": df_right_row[column].values[0],
                        "ConcatenatedKeys": keys,
                    }
                )

    return result_dict

def find_mismatched_rows(df_left, df_right):
    """
    Compare two dataframes based on the 'ConcatenatedKeys' column and return rows where 'HashValue' is not matching.

    Parameters:
    df_left (pandas.DataFrame): First dataframe.
    df_right (pandas.DataFrame): Second dataframe.

    Returns:
    pandas.DataFrame: A dataframe containing the rows where 'HashValue' is not matching based on the 'ConcatenatedKeys' column.
    """
    merged_df = pd.merge(df_left, df_right, on="ConcatenatedKeys", suffixes=("_df_left", "_df_right"))
    ic(merged_df)
    mismatched_df = merged_df[merged_df["HashValue_df_left"] != merged_df["HashValue_df_right"]]
    ic(mismatched_df)
    return mismatched_df

async def highlight_positive_cells(excel_filename, worksheet_name, column_name):
    """
    Highlights cells with positive values in the specified column of an Excel worksheet.

    Args:
        excel_filename (str): Path to the Excel file.
        worksheet_name (str): Name of the worksheet.
        column_name (str): Name of the column to check for positive values.

    Returns:
        None
    """
    try:
        # Read the Excel file
        df = pd.read_excel(excel_filename, sheet_name=worksheet_name)

        # Get the column index based on the column name
        column_index = df.columns.get_loc(column_name)

        # Create a workbook object
        wb = openpyxl.load_workbook(excel_filename)

        # Select the specified worksheet
        ws = wb[worksheet_name]

        # Define the red fill style
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

        # Iterate through rows and highlight positive cells
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=column_index + 1, max_col=column_index + 1):
            cell_value = row[0].value
            if cell_value is not None and cell_value > 0:
                row[0].fill = red_fill

        # Save the modified workbook
        wb.save(excel_filename)
        print(f"Positive values in column '{column_name}' highlighted in red.")

    except Exception as e:
        print(f"Error: {str(e)}")


async def highlight_cells_with_value(excel_filename, target_value):
    """
    Highlights cells in an Excel file that match the specified value.

    Args:
        excel_filename (str): Path to the Excel file.
        target_value (str): The value to search for in the cells.

    Returns:
        None
    """
    try:
        # Create a workbook object
        wb = openpyxl.load_workbook(excel_filename)

        # Iterate through all sheets in the workbook
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Define the red fill style
            red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

            # Iterate through all cells in the sheet
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value == target_value:
                        cell.fill = red_fill

        # Save the modified workbook
        wb.save(excel_filename)
        print(f"Cells with value '{target_value}' highlighted in red.")

    except Exception as e:
        print(f"Error: {str(e)}")


async def make_header_background_grey(excel_filename):
    """
    Modifies the background color of the header row in an Excel file to grey (if header cell has text).

    Args:
        excel_filename (str): Path to the Excel file.

    Returns:
        None
    """
    try:
        # Create a workbook object
        wb = openpyxl.load_workbook(excel_filename)

        # Iterate through all sheets in the workbook
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Define the grey fill style
            grey_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

            # Iterate through header cells (first row)
            for cell in ws[1]:
                if cell.value:
                    cell.fill = grey_fill

        # Save the modified workbook
        wb.save(excel_filename)
        print("Header background color set to grey for cells with text.")

    except Exception as e:
        print(f"Error: {str(e)}")
